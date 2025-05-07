"""Python implementation for 1 project, all test activities"""
import os
import pandas as pd
import math
import json
import xml.etree.ElementTree as ET
from datetime import datetime
import sys
import os.path
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy
import re

main_dict = {}

def shift_formula(formula, row_offset):
    def replacer(match):
        col = match.group(1)
        row = int(match.group(2)) + row_offset
        return f"{col}{row}"

    return re.sub(r"([A-Z]+)(\d+)", replacer, formula)

def append_from_json(gr_dir, project):
    json_file = os.path.join(gr_dir, project + "_01_review_test_spec.json")
    with open(json_file, 'r') as file:
        data = json.load(file)

    for key in data:
        module = key
        if module not in main_dict.keys():
            main_dict[module] = data[key]

def append_from_txt(gr_dir, project):
    txt_file = os.path.join(gr_dir, project + "_02_review_test_impl.txt")
    f = open(txt_file, "r")
    for line in f:
        module = line.split(':')[0]
        if module not in main_dict.keys():
            main_dict[module] = []
        results = line.split(':')[1]
        results = results.replace(" ", "")
        results = results.replace("\n", "")
        split_res = results.split(',')
        for res in split_res:
            main_dict[module].append(int(res))

def append_from_xml(gr_dir, project):
    xml_file = os.path.join(gr_dir, project + "_03_nonsafety_test.xml")
    tree = ET.parse(xml_file)
    modules = tree.findall("module")

    for module in modules:
        mod_name = module.find('name').text

        if mod_name not in main_dict.keys():
            main_dict[mod_name] = []

        results = module.find('results').text
        results = results.replace(" ", "")
        each_res = results.split(",")
        for res in each_res:
            main_dict[mod_name].append(int(res))

def append_from_xls(gr_dir, project):
    xls_file = os.path.join(gr_dir, project + "_04_safety_test.xlsx")
    data = pd.read_excel(xls_file, header=None, index_col=None)

    for row in data.iterrows():
        for i, elem in enumerate(row[1]):
            if i == 0:
                person = elem
                if person not in main_dict.keys():
                    main_dict[person] = []
            else:
                if not(math.isnan(elem)):
                    main_dict[person].append(int(elem))

def create_excel(gr_dir, project):
    py_dir = os.path.dirname(os.path.abspath(__file__))
    tst_dir = os.path.join(py_dir, "test_results")
    tpsr_path = os.path.join(tst_dir, "output.xlsx")

    wb = load_workbook(tpsr_path)
    ws = wb['Sheet1']

    template_row = 3
    no_rows = len(main_dict.keys()) - 1
    no_columns = 35

    for i in range(no_rows):
        target_row = template_row + i + 1
        for col in range(no_columns):
            source_cell = ws.cell(row=template_row, column=col+1)
            target_cell = ws.cell(row=template_row + i + 1, column=col+1)

            # Copy value or formula
            original_formula = source_cell.value
            if isinstance(original_formula, str) and original_formula.startswith("="):
                target_cell.value = shift_formula(original_formula, target_row - template_row)
            else:
                target_cell.value = source_cell.value

            target_cell._style = copy.copy(source_cell._style)
            target_cell.number_format = source_cell.number_format


    for i, each_row in enumerate(main_dict.items()):
        ws.cell(row=i+3, column=1).value = str(each_row[0])

    for m, key in enumerate(main_dict.keys()):
        print(main_dict[key])
        for n,elem in enumerate(main_dict[key]):
            if n <= 2:
                ws.cell(row=m+3, column=n+2).value = elem
            elif (n > 2 and n <= 5):
                ws.cell(row=m+3, column=n+5).value = elem
            elif (n > 5 and n <= 8):
                ws.cell(row=m+3, column=n+8).value = elem
            elif (n > 8 and n <= 11):
                ws.cell(row=m+3, column=n+10).value = elem
            elif (n > 11 and n <= 14):
                ws.cell(row=m+3, column=n+13).value = elem
            elif (n > 14 and n <= 17):
                ws.cell(row=m+3, column=n+15).value = elem
    wb.save(tpsr_path)

    print("Excel updated successfully")

def create_report(project):
    py_dir = os.path.dirname(os.path.abspath(__file__))
    gr_dir = os.path.join(py_dir, "test_results")

    append_from_json(gr_dir, project)
    append_from_txt(gr_dir, project)
    append_from_xml(gr_dir, project)
    append_from_xls(gr_dir, project)

    create_excel(gr_dir, project)

def main():
    arg = sys.argv
    if arg[1] == 'none':
        print("please select a valid option from build parameters")
    else:
        print("calling python script with project: " + arg[1])

    create_report(arg[1])

if __name__ == "__main__":
    main()